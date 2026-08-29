# -*- coding: utf-8 -*-
"""Lấy doanh thu theo ngày từ Pancake POS, lọc theo bộ phận, xuất Excel.

Chạy:
    python doanh_thu.py                              # hôm nay, tất cả nhân viên
    python doanh_thu.py 2026-08-28                   # ngày cụ thể
    python doanh_thu.py 2026-08-28 --bophan sale     # chỉ nhân viên thuộc bộ phận có chữ "sale"
    python doanh_thu.py 2026-08-01 2026-08-28 --bophan sale   # khoảng ngày
    ... --full                                       # hiện cả nhân viên 0 đơn

Cách tính (giống báo cáo Doanh thu -> Nhân viên trên POS):
    - ĐƠN CHỐT  = đơn có trạng thái: Đã xác nhận, Đã gửi hàng, Đã nhận,
                  Đang đóng hàng, Chờ chuyển hàng, Hoàn 1 phần, Đã thu tiền
    - DOANH THU = tổng tiền các đơn chốt
    - ĐƠN HOÀN  = đơn Đang hoàn / Đã hoàn; DOANH THU HOÀN = tổng tiền đơn hoàn
    - DOANH SỐ  = Doanh thu + Doanh thu hoàn
    - Tiền hàng = Tổng tiền - Phí VC thu của khách - Phụ thu
"""
import argparse
import sys
from datetime import date, datetime, time, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from pancake_client import PancakeClient, PancakeError

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

OUTPUT_DIR = Path(__file__).resolve().parent / "output"


def vnd(amount) -> str:
    try:
        return f"{int(amount):,.0f} đ".replace(",", ".")
    except (TypeError, ValueError):
        return "0 đ"


# ----------------------------------------------------------------------
# Nhân viên & bộ phận
# ----------------------------------------------------------------------
def clean_name(name: str) -> str:
    """Chuẩn hóa tên: bỏ khoảng trắng thừa để so khớp nhất quán."""
    return " ".join((name or "").split())


def load_staff(client: PancakeClient, shop_id: str) -> dict[str, dict]:
    """Trả về map: user_id -> {"name": tên, "dept": tên bộ phận}.

    Đồng thời lưu danh sách nhân viên thô vào api_data/nhanvien.json."""
    import luu_tru

    users = client.get_users(shop_id)
    luu_tru.luu_nhan_vien(users)
    staff = {}
    for u in users:
        uid = u.get("user_id") or (u.get("user") or {}).get("id")
        name = clean_name((u.get("user") or {}).get("name") or "")
        dept = (u.get("department") or {}).get("name") or "(chưa gán bộ phận)"
        if uid:
            staff[uid] = {"name": name or "(không rõ)", "dept": dept}
    return staff


def matched_departments(staff: dict[str, dict], keyword: str) -> set[str]:
    """Các bộ phận có tên chứa keyword (không phân biệt hoa/thường)."""
    kw = keyword.lower()
    return {s["dept"] for s in staff.values() if kw in s["dept"].lower()}


def seller_of(order: dict) -> tuple[str, str]:
    """(user_id, tên) của nhân viên phụ trách đơn."""
    seller = order.get("assigning_seller") or {}
    if seller.get("id") or seller.get("name"):
        return seller.get("id") or "", seller.get("name") or "(không rõ)"
    creator = order.get("creator") or {}
    return creator.get("id") or "", creator.get("name") or "(không rõ)"


def _blank_seller(name: str, dept: str) -> dict:
    return {
        "name": name, "dept": dept,
        "don_chot": 0, "doanh_thu": 0,
        "don_hoan": 0, "doanh_thu_hoan": 0,
        "tien_hang": 0, "phi_vc": 0, "phu_thu": 0,
    }


# ----------------------------------------------------------------------
# Lấy & tổng hợp dữ liệu 1 ngày
# ----------------------------------------------------------------------
def fetch_day(
    client: PancakeClient,
    shop_id: str,
    day: date,
    tz: ZoneInfo,
    staff: dict[str, dict],
    dept_filter: set[str] | None,
) -> dict:
    start_ts = int(datetime.combine(day, time.min, tzinfo=tz).timestamp())
    end_ts = int(datetime.combine(day, time.max, tzinfo=tz).timestamp())

    rows = []
    total_orders = 0
    revenue_all = 0        # doanh thu toàn shop (đơn chốt, tham khảo)
    by_status: dict[str, int] = {}
    summary = _blank_seller("TỔNG", "")   # cộng dồn theo bộ lọc

    # Khởi tạo sẵn MỌI nhân viên thuộc bộ phận được lọc, kể cả người không có đơn
    by_seller: dict[str, dict] = {}
    if dept_filter is not None:
        for uid, s in staff.items():
            if s["dept"] in dept_filter:
                by_seller[uid] = _blank_seller(s["name"], s["dept"])

    for o in client.iter_orders(shop_id, start_ts, end_ts, page_size=config.ORDERS_PAGE_SIZE):
        total_orders += 1
        status_code = o.get("status")
        status = config.ORDER_STATUS.get(status_code, f"Mã {status_code}")
        by_status[status] = by_status.get(status, 0) + 1

        total_price = o.get("total_price") or 0
        phi_vc = o.get("shipping_fee") or 0
        phu_thu = o.get("surcharge") or 0
        tien_hang = total_price - phi_vc - phu_thu

        if status_code in config.CLOSED_STATUSES:
            kind = "chot"
        elif status_code in config.RETURN_STATUSES:
            kind = "hoan"
        else:
            kind = None

        uid, name = seller_of(o)
        name = clean_name(name) or "(không rõ)"
        info = staff.get(uid)
        dept = info["dept"] if info else "(không rõ)"
        in_filter = dept_filter is None or dept in dept_filter

        if kind == "chot":
            revenue_all += total_price

        if kind and in_filter:
            key = uid or name
            s = by_seller.setdefault(key, _blank_seller(name, dept))
            for target in (s, summary):
                if kind == "chot":
                    target["don_chot"] += 1
                    target["doanh_thu"] += total_price
                    target["tien_hang"] += tien_hang
                    target["phi_vc"] += phi_vc
                    target["phu_thu"] += phu_thu
                else:
                    target["don_hoan"] += 1
                    target["doanh_thu_hoan"] += total_price

        if kind == "chot":
            phan_loai = "Đơn chốt" if in_filter else "Đơn chốt (ngoài bộ phận lọc)"
        elif kind == "hoan":
            phan_loai = "Đơn hoàn" if in_filter else "Đơn hoàn (ngoài bộ phận lọc)"
        else:
            phan_loai = f"Không tính ({status})"

        rows.append({
            "ma_don": o.get("id"),
            "thoi_gian_tao": o.get("inserted_at") or "",
            "trang_thai": status,
            "phan_loai": phan_loai,
            "tong_tien": total_price,
            "tien_hang": tien_hang,
            "phi_vc": phi_vc,
            "phu_thu": phu_thu,
            "cod": o.get("cod") or 0,
            "giam_gia": o.get("total_discount") or 0,
            "nhan_vien": name,
            "bo_phan": dept,
            "khach_hang": (o.get("customer") or {}).get("name") or o.get("bill_full_name") or "",
        })

    return {
        "day": day,
        "orders": rows,
        "summary": summary,
        "revenue_all": revenue_all,
        "total_orders": total_orders,
        "by_status": by_status,
        "by_seller": by_seller,
    }


def doanh_so(s: dict) -> int:
    return s["doanh_thu"] + s["doanh_thu_hoan"]


def ty_le_hoan(s: dict) -> float:
    """Tỷ lệ hoàn = đơn hoàn / (đơn chốt + đơn hoàn)."""
    n = s["don_chot"] + s["don_hoan"]
    return s["don_hoan"] / n if n else 0.0


def visible_sellers(result: dict, show_all: bool) -> list[dict]:
    """Danh sách NV hiển thị: mặc định chỉ người có đơn (giống báo cáo POS)."""
    sellers = [s for s in result["by_seller"].values()
               if show_all or s["don_chot"] > 0 or s["don_hoan"] > 0]
    return sorted(sellers, key=lambda x: (-x["doanh_thu"], x["name"]))


# ----------------------------------------------------------------------
# In báo cáo ra màn hình
# ----------------------------------------------------------------------
def print_report(result: dict, dept_filter: set[str] | None, show_all: bool = False) -> None:
    day = result["day"]
    t = result["summary"]
    print("=" * 62)
    print(f"DOANH THU NGÀY {day.strftime('%d/%m/%Y')}")
    if dept_filter is not None:
        print(f"Bộ phận lọc: {', '.join(sorted(dept_filter))}")
    print("Đơn chốt = Đã xác nhận / Đã gửi hàng / Đã nhận / Đang đóng hàng / "
          "Chờ chuyển hàng / Hoàn 1 phần / Đã thu tiền")
    print("=" * 62)
    print(f"Tổng số đơn trong ngày:  {result['total_orders']}")
    print(f"Đơn chốt:                {t['don_chot']}")
    print(f"DOANH THU:               {vnd(t['doanh_thu'])}")
    print(f"Đơn hoàn:                {t['don_hoan']}  (doanh thu hoàn: {vnd(t['doanh_thu_hoan'])})")
    print(f"Doanh số:                {vnd(doanh_so(t))}")
    if dept_filter is not None:
        print(f"(Doanh thu toàn shop:    {vnd(result['revenue_all'])})")

    sellers = visible_sellers(result, show_all)
    if sellers:
        print("\nTheo nhân viên:")
        print(f"  {'Nhân viên':<32} {'Chốt':>4} {'Doanh thu':>14} {'Hoàn':>4} {'DT hoàn':>12}")
        for s in sellers:
            print(f"  {s['name']:<32} {s['don_chot']:>4} {vnd(s['doanh_thu']):>14} "
                  f"{s['don_hoan']:>4} {vnd(s['doanh_thu_hoan']):>12}")
    if not show_all and dept_filter is not None:
        print("\n(Chỉ hiện nhân viên có đơn, giống báo cáo POS. Thêm --full để hiện cả nhân viên 0 đơn.)")
    print()


# ----------------------------------------------------------------------
# Xuất Excel
# ----------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F7A4D")
HEADER_FONT = Font(bold=True, color="FFFFFF")
MONEY_FMT = "#,##0"
PCT_FMT = "0.0%"

SUMMARY_COLS = [
    "Nhân viên", "Bộ phận", "Đơn chốt", "Doanh số", "Doanh thu",
    "Đơn hoàn", "Tỷ lệ hoàn", "Doanh thu hoàn",
    "Tiền hàng", "Phí VC thu của khách", "Phụ thu",
]
SUMMARY_WIDTHS = (35, 14, 10, 15, 15, 10, 10, 15, 15, 18, 12)

DETAIL_COLS = [
    ("Mã đơn", "ma_don", 10),
    ("Thời gian tạo", "thoi_gian_tao", 20),
    ("Trạng thái", "trang_thai", 16),
    ("Phân loại", "phan_loai", 26),
    ("Tổng tiền", "tong_tien", 14),
    ("Tiền hàng", "tien_hang", 14),
    ("Phí VC thu khách", "phi_vc", 15),
    ("Phụ thu", "phu_thu", 12),
    ("COD", "cod", 14),
    ("Giảm giá", "giam_gia", 12),
    ("Nhân viên", "nhan_vien", 32),
    ("Bộ phận", "bo_phan", 16),
    ("Khách hàng", "khach_hang", 24),
]


def _write_header(ws, titles):
    for col, title in enumerate(titles, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _write_summary_row(ws, row: int, s: dict, name: str | None = None, bold: bool = False):
    values = [
        name if name is not None else s["name"], s["dept"],
        s["don_chot"], doanh_so(s), s["doanh_thu"],
        s["don_hoan"], ty_le_hoan(s), s["doanh_thu_hoan"],
        s["tien_hang"], s["phi_vc"], s["phu_thu"],
    ]
    money_cols = {4, 5, 8, 9, 10, 11}
    for col, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=v)
        if col in money_cols:
            cell.number_format = MONEY_FMT
        elif col == 7:
            cell.number_format = PCT_FMT
        if bold:
            cell.font = Font(bold=True)


def export_excel(result: dict, dept_filter: set[str] | None, show_all: bool = False) -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    suffix = "_locbophan" if dept_filter is not None else ""
    path = OUTPUT_DIR / f"doanhthu_{result['day'].isoformat()}{suffix}.xlsx"

    wb = Workbook()

    # --- Sheet 1: Tổng hợp theo nhân viên (đủ cột như báo cáo POS) ---
    ws = wb.active
    ws.title = "Tổng hợp NV"
    _write_header(ws, SUMMARY_COLS)
    row = 2
    for s in visible_sellers(result, show_all):
        _write_summary_row(ws, row, s)
        row += 1
    _write_summary_row(ws, row, result["summary"], name="TỔNG", bold=True)
    for col, width in enumerate(SUMMARY_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # --- Sheet 2: Chi tiết đơn ---
    ws2 = wb.create_sheet("Chi tiết đơn")
    _write_header(ws2, [t for t, _, _ in DETAIL_COLS])
    for r, order in enumerate(result["orders"], start=2):
        for col, (_, key, _) in enumerate(DETAIL_COLS, start=1):
            cell = ws2.cell(row=r, column=col, value=order[key])
            if key in ("tong_tien", "tien_hang", "phi_vc", "phu_thu", "cod", "giam_gia"):
                cell.number_format = MONEY_FMT
    for col, (_, _, width) in enumerate(DETAIL_COLS, start=1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_COLS))}{len(result['orders']) + 1}"

    try:
        wb.save(path)
    except PermissionError:
        # File đang mở trong Excel -> lưu sang tên khác kèm giờ phút
        stamp = datetime.now().strftime("%H%M%S")
        path = path.with_name(path.stem + f"_{stamp}" + path.suffix)
        wb.save(path)
        print("(File cũ đang mở trong Excel nên đã lưu sang tên mới. "
              "Đóng file Excel trước khi chạy để ghi đè file cũ.)")
    return path


def parse_args() -> tuple[date, date, str | None, bool]:
    parser = argparse.ArgumentParser(description="Doanh thu theo ngày từ Pancake POS")
    parser.add_argument("start", nargs="?", help="Ngày bắt đầu YYYY-MM-DD (mặc định: hôm nay)")
    parser.add_argument("end", nargs="?", help="Ngày kết thúc YYYY-MM-DD (mặc định: bằng ngày bắt đầu)")
    parser.add_argument("--bophan", metavar="TUKHOA", default=None,
                        help='Chỉ tính nhân viên thuộc bộ phận có tên chứa từ khóa này, ví dụ: --bophan sale')
    parser.add_argument("--full", action="store_true",
                        help="Hiện cả nhân viên 0 đơn của bộ phận lọc (mặc định: chỉ hiện người có đơn, giống POS)")
    args = parser.parse_args()

    tz = ZoneInfo(config.TIMEZONE)
    today = datetime.now(tz).date()
    try:
        start_day = date.fromisoformat(args.start) if args.start else today
        end_day = date.fromisoformat(args.end) if args.end else start_day
    except ValueError:
        raise SystemExit("Ngày không hợp lệ. Dùng định dạng YYYY-MM-DD, ví dụ: python doanh_thu.py 2026-08-28")
    return start_day, end_day, args.bophan, args.full


def main() -> None:
    api_key = config.require_api_key()
    shop_id = config.PANCAKE_SHOP_ID
    if not shop_id:
        raise SystemExit("Chưa có PANCAKE_SHOP_ID trong .env. Chạy 'python test_connection.py' để xem danh sách shop_id.")

    start_day, end_day, dept_keyword, show_all = parse_args()
    tz = ZoneInfo(config.TIMEZONE)
    client = PancakeClient(api_key)

    staff = load_staff(client, shop_id)

    dept_filter: set[str] | None = None
    if dept_keyword:
        dept_filter = matched_departments(staff, dept_keyword)
        if not dept_filter:
            all_depts = ", ".join(sorted({s["dept"] for s in staff.values()}))
            raise SystemExit(
                f"Không có bộ phận nào chứa '{dept_keyword}'.\nCác bộ phận hiện có: {all_depts}"
            )
        print(f"Bộ phận khớp '{dept_keyword}': {', '.join(sorted(dept_filter))}\n")

    grand_total = 0
    day = start_day
    while day <= end_day:
        try:
            result = fetch_day(client, shop_id, day, tz, staff, dept_filter)
        except PancakeError as e:
            raise SystemExit(f"Lỗi khi lấy đơn ngày {day}: {e}")
        print_report(result, dept_filter, show_all)
        xlsx_path = export_excel(result, dept_filter, show_all)
        print(f"Đã xuất Excel: {xlsx_path}\n")
        grand_total += result["summary"]["doanh_thu"]
        day += timedelta(days=1)

    if start_day != end_day:
        print("=" * 62)
        print(f"TỔNG DOANH THU {start_day.strftime('%d/%m')} - {end_day.strftime('%d/%m/%Y')}: {vnd(grand_total)}")


if __name__ == "__main__":
    main()
